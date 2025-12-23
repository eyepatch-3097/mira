# sources/services/sheets.py
import csv
import os
from openpyxl import load_workbook

def _cell(v):
    if v is None:
        return ""
    return str(v)

def preview_xlsx(path: str, max_rows: int = 10):
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        headers = []
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [_cell(x) for x in row]
                # fallback header names
                headers = [h if h.strip() else f"Column {idx+1}" for idx, h in enumerate(headers)]
                continue
            if i > max_rows:
                break
            rows.append([_cell(x) for x in row])
        out.append({
            "sheet_name": ws.title,
            "headers": headers,
            "rows": rows,
        })
    return out

def preview_csv(path: str, max_rows: int = 10):
    # try utf-8 then fallback
    for enc in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            with open(path, "r", newline="", encoding=enc) as f:
                reader = csv.reader(f)
                headers = next(reader, [])
                headers = [h.strip() if h.strip() else f"Column {i+1}" for i, h in enumerate(headers)]
                rows = []
                for i, row in enumerate(reader):
                    if i >= max_rows:
                        break
                    rows.append([c for c in row])
                return {"headers": headers, "rows": rows}
        except Exception:
            continue
    return {"headers": [], "rows": []}
